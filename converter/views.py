from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .utils import convert_pdf_to_excel
from .models import UploadedPDF
import os
from rest_framework import generics
from .models import UploadedPDF
from .serializer import UploadedPDFSerializer

class UploadedPDFListView(generics.ListAPIView):
    queryset = UploadedPDF.objects.all()
    serializer_class = UploadedPDFSerializer


@csrf_exempt
def hello_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)

    he_dao_tao = request.POST.get('he_dao_tao')
    danh_sach_quyet_dinh = request.POST.get('danh_sach_quyet_dinh')

    if not he_dao_tao:
        return JsonResponse({'error': 'Thiếu tham số hệ đào tạo'}, status=400)
    if not danh_sach_quyet_dinh:
        return JsonResponse({'error': 'Thiếu danh sách quyết định'}, status=400)

    files = request.FILES.getlist('pdf_files')
    if not files:
        return JsonResponse({'error': 'No PDF files uploaded'}, status=400)

    output_dir = 'converted_excels'
    os.makedirs(output_dir, exist_ok=True)

    result = []

    for uploaded_file in files:
        # Lưu vào model
        saved_record = UploadedPDF.objects.create(
            pdf_file=uploaded_file,
            he_dao_tao=he_dao_tao,
            danh_sach_quyet_dinh=danh_sach_quyet_dinh
        )

        wb, preview_data, extra_info = convert_pdf_to_excel(uploaded_file)
        filename = uploaded_file.name.replace('.pdf', '.xlsx')
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)

        result.append({
            'filename': filename,
            'preview_data': preview_data,
            'so': extra_info.get("so"),
            'ngay_thang_nam': extra_info.get("ngay_thang_nam"),
            'id_saved': saved_record.id
        })

    return JsonResponse({
        'status': '200',
        'he_dao_tao': he_dao_tao,
        'danh_sach_quyet_dinh': danh_sach_quyet_dinh,
        'results': result
    })


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import UploadedPDF
import os


@csrf_exempt
def delete_uploaded_pdfs(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)

    # Lấy tham số 'he_dao_tao' và 'danh_sach_quyet_dinh' từ body param
    he_dao_tao = request.POST.get('he_dao_tao')
    # danh_sach_quyet_dinh = request.POST.get('danh_sach_quyet_dinh')

    if not he_dao_tao:
        return JsonResponse({'error': 'Thiếu tham số he_dao_tao hoặc danh_sach_quyet_dinh'}, status=400)

    # Lọc dữ liệu theo các tham số 'he_dao_tao' và 'danh_sach_quyet_dinh'
    files_to_delete = UploadedPDF.objects.filter(
        he_dao_tao=he_dao_tao,
        # danh_sach_quyet_dinh=danh_sach_quyet_dinh
    )

    # Nếu không có dữ liệu nào để xóa
    if not files_to_delete.exists():
        return JsonResponse({'error': 'No matching records found'}, status=404)

    # Xóa từng file và bản ghi
    for pdf in files_to_delete:
        # Xóa file PDF trong hệ thống (nếu tồn tại)
        if os.path.exists(pdf.pdf_file.path):
            os.remove(pdf.pdf_file.path)

        # Xóa bản ghi trong cơ sở dữ liệu
        pdf.delete()

    return JsonResponse({'status': '200', 'message': 'Matching PDFs have been deleted'})

import openpyxl
import os
import requests
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings

@csrf_exempt
def upload_excel_api(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)

    he_dao_tao = request.POST.get('he_dao_tao')
    danh_sach_quyet_dinh = request.POST.get('danh_sach_quyet_dinh')
    excel_file = request.FILES.get('excel_file')
    pdf_files = request.FILES.getlist('pdf_files')

    if not he_dao_tao or not danh_sach_quyet_dinh or not excel_file:
        return JsonResponse({'error': 'Thiếu dữ liệu đầu vào'}, status=400)

    # Gửi dữ liệu đến hello_api
    try:
        hello_response = requests.post(
            'http://172.31.65.71:8000/api/hello/',
            data={'he_dao_tao': he_dao_tao, 'danh_sach_quyet_dinh': danh_sach_quyet_dinh},
            files=[('pdf_files', (f.name, f, f.content_type)) for f in pdf_files]
        )
        hello_data = hello_response.json()
        results = hello_data.get('results', [])

    except Exception as e:
        return JsonResponse({'error': f'Lỗi khi gọi hello_api: {str(e)}'}, status=500)

    # Tạo dict dữ liệu từ hello_api để tra cứu
    api_students = {}
    for item in results:
        for student in item.get('preview_data', []):
            mssv = student.get('mssv')
            api_students[mssv] = student

    # Đọc file Excel
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Cột kết quả sẽ được cập nhật
    col_mapping = {
        'Quyết định công nhận sinh viên': 15,  # Cột P
        'Quyết định công nhận chuyên ngành': 16,  # Cột Q
        'Quyết định miễn giảm': 17,  # Cột R
        'Quyết định chuyển ngành': 18,  # Cột S
        'Quyết định bảo lưu': 19,  # Cột T
        'Quyết định thôi học': 20,  # Cột U
        'Quyết định chuyển cơ sở': 21,  # Cột V
        'Quyết định khen thưởng': 22,  # Cột W
        'Quyết định kỷ luật': 23,  # Cột X
        'Quyết định tốt nghiệp Trung cấp': 24  # Cột Y
    }

    so_dong_cap_nhat = 0
    so_dong_doi_chieu = 0

    for row in ws.iter_rows(min_row=2):
        mssv_cell = row[1]   # Cột B - MSSV
        hoten_cell = row[2]  # Cột C - Họ tên SV
        ngaysinh_cell = row[14]  # Cột 0
        gioitinh_cell = row[15]  # Cột P
        dantoc_cell = row[16]  # Cột Q
        ghichu_cell = row[1]  # Cột AB (26th column, index 25)

        if he_dao_tao == 'FPT Polytechnic':
            ngaysinh_cell = row[14]  # Cột 0
            gioitinh_cell = row[15]  # Cột P
            dantoc_cell = row[16]  # Cột Q
            ghichu_cell = row[27]  # Cột AB (26th column, index 25)

        if he_dao_tao == 'FPT Polyschool':
            ngaysinh_cell = row[12]  # Cột M
            gioitinh_cell = row[13]  # Cột N
            dantoc_cell = row[14]    # Cột O
            ghichu_cell = row[25]    # Cột Z (26th column, index 25)

        mssv = str(mssv_cell.value).strip() if mssv_cell.value else None
        if mssv and mssv in api_students:
            so_dong_doi_chieu += 1
            student_data = api_students[mssv]

            note_parts = []

            # Họ tên
            excel_name = str(hoten_cell.value).strip() if hoten_cell.value else ""
            api_name = str(student_data.get('họ và tên', '')).strip()
            if excel_name and excel_name.lower() != api_name.lower():
                note_parts.append(f"Họ tên khác: '{excel_name}' ≠ '{api_name}'")

            # Ngày sinh
            excel_dob = str(ngaysinh_cell.value).strip() if ngaysinh_cell.value else ""
            api_dob = str(student_data.get('ngày sinh', '')).strip()
            if excel_dob and excel_dob != api_dob:
                note_parts.append(f"Ngày sinh khác: '{excel_dob}' ≠ '{api_dob}'")
            elif not excel_dob and api_dob:
                ngaysinh_cell.value = api_dob
                so_dong_cap_nhat += 1

            # Giới tính
            excel_gt = str(gioitinh_cell.value).strip() if gioitinh_cell.value else ""
            api_gt = str(student_data.get('giới tính', '')).strip()
            if excel_gt and excel_gt.lower() != api_gt.lower():
                note_parts.append(f"Giới tính khác: '{excel_gt}' ≠ '{api_gt}'")
            elif not excel_gt and api_gt:
                gioitinh_cell.value = api_gt
                so_dong_cap_nhat += 1

            # Dân tộc
            excel_dt = str(dantoc_cell.value).strip() if dantoc_cell.value else ""
            api_dt = str(student_data.get('dân tộc', '')).strip()
            if excel_dt and excel_dt.lower() != api_dt.lower():
                note_parts.append(f"Dân tộc khác: '{excel_dt}' ≠ '{api_dt}'")
            elif not excel_dt and api_dt:
                dantoc_cell.value = api_dt
                so_dong_cap_nhat += 1

            if note_parts:
                ghichu_cell.value = "; ".join(note_parts)

            # Lấy các quyết định từ API
            # Khởi tạo dict chứa mssv -> so quyết định
            decision_mapping = {}

            for decision in results:
                so = decision.get('so', '')
                ngay = decision.get('ngay_thang_nam', '') or ''
                decision_text = f"{so} {ngay}".strip()



                for sv in decision.get('preview_data', []):
                    mssv = sv.get('mssv')
                    if mssv:
                        decision_mapping[mssv] = decision_text

            print('decision_mapping')
            print(decision_mapping)

            # "Quyết định công nhận sinh viên": 15,  # Cột P
            # "Quyết định công nhận chuyên ngành": 16,  # Cột Q
            # "Quyết định miễn giảm": 17,  # Cột R
            # "Quyết định chuyển ngành": 18,  # Cột S
            # "Quyết định bảo lưu": 19,  # Cột T
            # "Quyết định thôi học": 20,  # Cột U
            # "Quyết định chuyển cơ sở": 21,  # Cột V
            # "Quyết định khen thưởng": 22,  # Cột W
            # "Quyết định kỷ luật": 23,  # Cột X
            # "Quyết định tốt nghiệp Trung cấp": 24  # Cột Y

            # Cập nhật cột trong Excel tùy thuộc vào loại quyết định
            for row in ws.iter_rows(min_row=2):
                mssv_cell = row[1]  # Cột B (index 1)
                mssv = str(mssv_cell.value).strip() if mssv_cell.value else None

                if he_dao_tao == 'FPT Polytechnic':
                    if mssv and mssv in decision_mapping:
                        if danh_sach_quyet_dinh == 'Công nhận sinh viên':
                            row[17].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Chuyển ngành':
                            row[18].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Chuyển cơ sở':
                            row[19].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Nghỉ học tạm thời':
                            row[20].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Nhập học trở lại':
                            row[21].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Chuyển khung':
                            row[22].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Miễn giảm môn học':
                            row[23].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Khen thưởng':
                            row[24].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Kỷ luật':
                            row[25].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Công nhận tốt nghiệp':
                            row[26].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)

                if he_dao_tao == 'FPT Polyschool':
                    if mssv and mssv in decision_mapping:
                        if danh_sach_quyet_dinh == 'Quyết định công nhận sinh viên':
                            row[15].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định công nhận chuyên ngành':
                            row[16].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định miễn giảm':
                            row[17].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định chuyển ngành':
                            row[18].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định bảo lưu':
                            row[19].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định thôi học':
                            row[20].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định chuyển cơ sở':
                            row[21].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định khen thưởng':
                            row[22].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định kỷ luật':
                            row[23].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)
                        if danh_sach_quyet_dinh == 'Quyết định tốt nghiệp Trung cấp':
                            row[24].value = decision_mapping[mssv]  # Cột P là index 15 (đếm từ 0)


    # Lưu file mới
    output_dir = os.path.join(settings.MEDIA_ROOT, 'excel_outputs')
    os.makedirs(output_dir, exist_ok=True)
    output_filename = 'excel_doi_chieu.xlsx'
    output_path = os.path.join(output_dir, output_filename)
    wb.save(output_path)

    file_url = request.build_absolute_uri(settings.MEDIA_URL + f'excel_outputs/{output_filename}')
    return JsonResponse({
        'status': 'success',
        'so_luong_mssv_doi_chieu': so_dong_doi_chieu,
        'so_dong_duoc_cap_nhat': so_dong_cap_nhat,
        'file_output': file_url
    })
